import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side
from tkinter import filedialog, Tk, colorchooser
import os
import subprocess

def lighten_color(hex_color, percentage):
    hex_color = hex_color.lstrip("#")
    r, g, b = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
    r = min(int(r + (255 - r) * percentage / 100), 255)
    g = min(int(g + (255 - g) * percentage / 100), 255)
    b = min(int(b + (255 - b) * percentage / 100), 255)
    return f"#{r:02X}{g:02X}{b:02X}"

def choose_color(title):
    color_code = colorchooser.askcolor(title=title)[1]
    return color_code or "#FFFFFF"

def select_csv_file():
    root = Tk()
    root.withdraw()
    return filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv")])

def select_output_file():
    root = Tk()
    root.withdraw()
    return filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])

def beautify_csv_to_xlsx(csv_file, output_file, header_color, row_color):
    secondary_row_color = lighten_color(row_color, 80)
    df = pd.read_csv(csv_file)
    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name="Sheet1")
    workbook = writer.book
    sheet = workbook["Sheet1"]
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        sheet.column_dimensions[column].width = max_length + 2
    header_fill = PatternFill(start_color=header_color.lstrip("#"), end_color=header_color.lstrip("#"), fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column)):
        row_fill_color = row_color if i % 2 == 0 else secondary_row_color
        row_fill = PatternFill(start_color=row_fill_color.lstrip("#"), end_color=row_fill_color.lstrip("#"), fill_type="solid")
        for cell in row:
            cell.fill = row_fill
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = border
    writer.close()
    return output_file

def open_folder_and_highlight(file_path):
    folder_path = os.path.dirname(file_path)
    file_name = os.path.basename(file_path)
    if os.name == 'nt':
        folder_path = folder_path.replace("/", "\\")
        subprocess.run(['explorer', '/select,', os.path.join(folder_path, file_name)])
    else:
        subprocess.run(['open', folder_path])

csv_file = select_csv_file()
if csv_file:
    output_file = select_output_file()
    if output_file:
        header_color = choose_color("Choose a Header Color")
        row_color = choose_color("Choose a Row Color")
        beautified_xlsx = beautify_csv_to_xlsx(csv_file, output_file, header_color, row_color)
        print(f"Beautified file saved as: {beautified_xlsx}")
        open_folder_and_highlight(beautified_xlsx)
